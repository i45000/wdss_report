from openpyxl import Workbook
from datetime import datetime, timedelta

import requests
from bs4 import BeautifulSoup

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the title for column A
ws['A1'] = 'Maturity'
ws['B1'] = '1 Month'
ws['C1'] = '3 Months'
ws['D1'] = '6 Months'
ws['E1'] = '12 Months'

# Extract the data for one month and three months
one_month = None
three_months = None
six_months = None
twelve_months = None


def get_hibor_rate():
    if table is not None:
        global one_month, three_months, six_months, twelve_months

        #Finad the 
        tbody = table.find_all('tr')
        #print(tbody)
        # Define a list of time periods to extract rates for
        time_periods = ['1 Month', '3 Months','6 Months','12 Months']

        # Extract the rates for each time period
        for row in tbody:
            td_tags = row.find_all('td', {'align': True})
            #print(td_tags)
            #print("this is  len ",len(td_tags))
            if len(td_tags) == 2:
                time_period = td_tags[0].get_text().strip()
                #print(time_period)
                rate = td_tags[1].get_text().strip()
                if time_period in time_periods:
                    if time_period == '1 Month':
                        one_month = float(rate)
                    elif time_period == '3 Months':
                        three_months = float(rate)
                    elif time_period == '6 Months':
                        six_months =float(rate)
                    elif time_period == '12 Months':
                        twelve_months =float(rate)
        if (one_month is not None) or (twelve_months is not None):
            print('One month--B2:', one_month)
            print('Three months:--B3', three_months)
            print('Six months--B4:', six_months)
            print('Twelve months--B5:', twelve_months)
        else:
            print("non-working day")
    else:
        print('Table not found')


# Generate a list of all working days between two dates
start_date = datetime(2018, 1, 1)
end_date = datetime(2022, 12, 31)
dates = []
current_date = start_date
while current_date <= end_date:
    # Check if the current date is a working day (Monday to Friday)
    if current_date.weekday() not in [5, 6]:
        dates.append(current_date)
    current_date += timedelta(days=1)

# Add dates to column A
for i, date in enumerate(dates):
    # Format the date as YYYY-MM-DD
    date_str = date.strftime('%Y-%m-%d')
    url_hibor='https://www.hkab.org.hk/hibor/listRates.do?lang=en&Submit=Search&year='+str(date.year)+'&month='+str(date.month)+'&day='+str(date.day)
    response = requests.get(url_hibor)

    soup = BeautifulSoup(response.text, 'html.parser')

    # Find the table containing the data
    table = soup.find('table', {'class': 'etxtmed','bgcolor':'#ffffff'})
    get_hibor_rate()
    # Write the date to the cell in column A
    ws.cell(row=i+2, column=1).value = date_str
    ws.cell(row=i+2, column=2).value = one_month
    ws.cell(row=i+2, column=3).value = three_months
    ws.cell(row=i+2, column=4).value = six_months
    ws.cell(row=i+2, column=5).value = twelve_months

    one_month = None
    three_months = None
    six_months = None
    twelve_months = None


    

# Save the workbook
wb.save('C:\\Users\\jjcky\\Desktop\\finance\\example.xlsx')